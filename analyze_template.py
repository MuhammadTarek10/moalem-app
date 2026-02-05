import openpyxl
from openpyxl.utils import get_column_letter

# Load the workbook
file_path = 'assets/files/كشف فارغ اعمال السنة اولى وتانية ابتدائى.xlsx'
try:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")

    # Merged Cells
    print("\nMerged Cells:")
    for range_ in sheet.merged_cells.ranges:
        print(range_)

    # Column Dimensions
    print("\nColumn Dimensions:")
    for col in sheet.column_dimensions:
        dim = sheet.column_dimensions[col]
        print(f"Column {col}: Width={dim.width}")

    # Row Dimensions (First 20)
    print("\nRow Dimensions (First 20):")
    for row in range(1, 21):
        if row in sheet.row_dimensions:
             dim = sheet.row_dimensions[row]
             print(f"Row {row}: Height={dim.height}")
        else:
             print(f"Row {row}: Default Height")

    # Cell Details (First 15 Rows)
    print("\nCell Details (First 15 Rows):")
    for row in range(1, 16):
        for col in range(1, 30):  # Check first 30 columns
            cell = sheet.cell(row=row, column=col)
            val = cell.value
            if val or cell.has_style:
                col_letter = get_column_letter(col)
                print(f"Cell {col_letter}{row}:")
                if val:
                    print(f"  Value: {val}")
                
                # Font
                font = cell.font
                color_hex = "None"
                if font.color:
                     if hasattr(font.color, 'rgb'):
                         color_hex = str(font.color.rgb)
                     elif hasattr(font.color, 'theme'):
                         color_hex = f"Theme {font.color.theme}"
                print(f"  Font: Name={font.name}, Size={font.sz}, Bold={font.b}, Color={color_hex}")
                
                # Fill
                fill = cell.fill
                fill_color = "None"
                if hasattr(fill, 'fgColor'):
                     if getattr(fill.fgColor, 'rgb', None):
                         fill_color = str(fill.fgColor.rgb)
                     elif getattr(fill.fgColor, 'theme', None):
                         fill_color = f"Theme {fill.fgColor.theme}"
                
                if fill.fill_type:
                     print(f"  Fill: Type={fill.fill_type}, Color={fill_color}")
                
                # Border
                border = cell.border
                if border.left.style or border.right.style or border.top.style or border.bottom.style:
                    print(f"  Border: Left={border.left.style}, Right={border.right.style}, Top={border.top.style}, Bottom={border.bottom.style}")
                
                # Alignment
                align = cell.alignment
                print(f"  Align: H={align.horizontal}, V={align.vertical}, Wrap={align.wrap_text}, Rotation={align.text_rotation}")

except Exception as e:
    print(f"Error: {e}")
