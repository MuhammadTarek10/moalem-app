import openpyxl
import os

file_path = r"c:\Users\Lenovo\moalem-app\assets\files\كشف فارغ اعمال السنة اولى وتانية ابتدائى.xlsx"

print(f"\n{'='*50}")
print(f"Inspecting File: {os.path.basename(file_path)}")
print(f"{'='*50}")

try:
    if not os.path.exists(file_path):
        print(f"File not found at: {file_path}")
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Assuming the sheet with data is active or first, but let's check names
        print(f"Sheet Names: {wb.sheetnames}")
        
        sheet = wb.active
        print(f"Active Sheet: {sheet.title}")

        # Scan row 5, 6, 7 for headers and scores
        print("\n--- Headers (Rows 5-7) & Scores (Row 8) ---\n")
        for c in range(12, 20): # Focus on one week block
            header5 = str(sheet.cell(row=5, column=c).value).strip()
            header6 = str(sheet.cell(row=6, column=c).value).strip()
            header7 = str(sheet.cell(row=7, column=c).value).strip()
            score = str(sheet.cell(row=8, column=c).value).strip()
            print(f"Col {c}: R5='{header5}' R6='{header6}' R7='{header7}' (Max: {score})")

except Exception as e:
    print(f"Error inspecting {file_path}: {e}")
