import openpyxl
import sys

file_path = "c:\\Users\\Lenovo\\moalem-app\\كشف فارغ الغياب.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"Sheet Title: {ws.title}")
    print(f"Max Row: {ws.max_row}")
    print(f"Max Column: {ws.max_column}")
    
    print("\nMerged Cells:")
    for range_ in ws.merged_cells.ranges:
        print(range_)
        
    print("\nContent Preview (First 15 rows):")
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=10):
        row_data = []
        for cell in row:
            val = cell.value
            if val is None:
                val = ""
            else:
                val = str(val).strip()
            row_data.append(val)
        print("\t".join(row_data))
        
except ImportError:
    print("Error: openpyxl is not installed. Please install it using 'pip install openpyxl'")
except Exception as e:
    print(f"Error: {e}")
